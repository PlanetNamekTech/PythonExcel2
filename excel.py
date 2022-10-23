import openpyxl
from openpyxl import Workbook, load_workbook

wb = load_workbook('C2-W3-Practice-Challenge.xlsx')
ws = wb.active
#Looks at the value of the cell
print(ws['C15'].value)
#If a change is made like below, the workbook needs to be saved
ws['B2'].value = "Test" #can assign a value without using .value
wb.save('C2-W3-Practice-Challenge.xlsx')
#To create a new sheet
wb.create_sheet('PythonSheet')
wb.save('C2-W3-Practice-Challenge.xlsx')
#show names of all sheets in the workbook
print(wb.sheetnames)




