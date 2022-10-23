from json import load
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# to start a new workbook --> wb = Workbook()
wb = load_workbook('tacos.xlsx')
ws = wb.active
# how to change the active sheet name --> ws.title = "Data"

for row in range(1,11):
  for col in range(1,5):
    char = get_column_letter(col)
    ws[char + str(row)] = char + str(row)
  
wb.save('tacos.xlsx')



# how to append rows - ws.append(['I', 'am','great'])
#To save the newly created workbook with a name
# to save the new workbook --> wb.save('tacos.xlsx')