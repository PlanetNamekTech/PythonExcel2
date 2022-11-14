from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('tacos.xlsx')
ws = wb.active

#Specify ranges to be merged
ws.merge_cells("A1:D1")
#Unmerge cells, will lose data previously merged
ws.unmerge_cells("A1:D1")
#Insert or delete an empty row after row 7
ws.insert_rows(7)
ws.delete_rows(7)
#Inserting and deleting columns is more or less the same (insert_cols)


wb.save('tacos.xlsx')